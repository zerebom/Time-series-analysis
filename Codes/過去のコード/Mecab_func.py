import openpyxl#xlsxの操作に用いる
import urllib#urlを指定の形に変更する
import untangle#xmlの読み込みに用いる
from lxml import etree#xmlの構造・要素を取得するのに用いられる
import urllib3
import pandas as pd
import xlsxwriter
import numpy as np
import MeCab
from collections import OrderedDict
from collections import Counter
import glob

def make_lines2(fname_parsed):#ジェネレーター
    # MeCabfileが必要↑

    with open(fname_parsed) as file_parsed:       
# ,'r',errors='ignore',encoding='utf-8_sig'
        morphemes = []
        
 
        for line in file_parsed:
       
            cols = line.split('\t')
            if(len(cols) < 2):
                raise StopIteration     # 区切りがなければ終了
            # ここでリスト型になる↓   
            res_cols = cols[1].split(',')

            # 辞書作成、リストに追加
    
            # cols[0]=re.sub('.+　','',cols[0])  
            if  cols[0]!='\u3000':           
                morphemes.append(cols[0])

            # 品詞細分類1が'句点'なら文の終わりと判定
            if res_cols[1] == '句点':

                # morphemes.append(cols[0])
                # ↑これがあると。が二つになってしまう
                yield morphemes
                # print(morphemes)
                morphemes = []

#あるフォルダに入っているtxtファイルを一つずつ読み込み、一つずつMeCabファイルに変換して保存する
def gather3(open_file):
    print('a')
    for file in glob.glob(open_file):
        print(file)
        to_mecab(file,'{}.mecab'.format(file))


def count_morpheme3(write_file,mecab_file):

    copus=[]
    #語数カウンター
    word=""
    for morphemes in make_lines2(mecab_file):
        if len(morphemes)>=5:
            copus.append(morphemes[-5]+morphemes[-4]+morphemes[-3]+morphemes[-2])
        elif len(morphemes)>=4:
            copus.append(morphemes[-4]+morphemes[-3]+morphemes[-2])
       
        elif len(morphemes)>=3:
    
            copus.append(morphemes[-3]+morphemes[-2])
       
        elif len(morphemes)>=2:
            copus.append(morphemes[-2])
       

    word=''.join(copus)
  




#とあるフォルダの中のファイルを一つずつ開き、一つにまとめて出力する。
def gather2(open_file):
    merge_speech=""
    speech_list=[]
    for file in glob.glob(open_file):
        # to_mecab(file,'{}.mecab'.format(file))
        with open(file,errors='igonre') as write_file:
            write_file=write_file.read()
            #この一文がないと、io.TextIOWrapperの形で出力される
            speech_list.append(write_file)
            merge_speech+=write_file

    return(speech_list)

#とあるフォルダのなかのファイルをひとつずつ開き、txtファイルにまとめて出力する
#to_mecabのコメントアウトを入れると、その他に、各ファイルをmecabファイルに変換し、保存する
def gather(open_file,write_file):
    merge_speech=""
    speech_list=[]
    for file in glob.glob(open_file):
        # to_mecab(file,'{}.mecab'.format(file))
        with open(file,errors='igonre') as write_file:
            write_file=write_file.read()
            #この一文がないと、io.TextIOWrapperの形で出力される
            speech_list.append(write_file)
            merge_speech+=write_file

    with open(write_file,mode='w',errors='ignore') as out_file:
        out_file.write(merge_speech)
        return(speech_list)

#txtファイルを引数に、mecabファイルを生成し、保存する
def to_mecab(fname,fname_parsed):
    # ,encoding='utf-8_sig'
    with open(fname,errors='ignore') as data_file, \
            open(fname_parsed, mode='w',errors='ignore') as out_file:
                mecab = MeCab.Tagger(r'-d C:\Users\icech\mecab-ipadic-neologd\build\mecab-ipadic-2.7.0-20070801-neologd-20180625')
                out_file.write(mecab.parse(data_file.read()))


        # mecab = MeCab.Tagger('-d /var/lib/mecab/dic/mecab-ipadic-neologd')

        

#雌株ファイルを引数にとり、一文ずつ（句点区切り）でよみこみ、形態素解析したものに対して、ラベルを付けて
# ジェネレータとして出力する
def make_lines(fname_parsed):#ジェネレーター
    '''
    各形態素を
    ・表層形（surface）
    ・基本形（base）
    ・品詞（pos）
    ・品詞細分類1（pos1）
    の4つをキーとする辞書に格納し、1文ずつ、この辞書のリストとして返す

    戻り値：
    1文の各形態素を辞書化したリスト
    '''
    with open(fname_parsed) as file_parsed:

        morphemes = []
        for line in file_parsed:

            # 表層形はtab区切り、それ以外は','区切りでバラす
            cols = line.split('\t')
            if(len(cols) < 2):
                raise StopIteration     # 区切りがなければ終了
            res_cols = cols[1].split(',')

            # 辞書作成、リストに追加
            morpheme = {
                'surface': cols[0],
                'base': res_cols[6],
                'pos': res_cols[0],
                'pos1': res_cols[1]
            }
            morphemes.append(morpheme)

            # 品詞細分類1が'句点'なら文の終わりと判定
            if res_cols[1] == '句点':
                yield morphemes
                morphemes = []

#make_lineを用いて、for文を回す。
# そして、目的の品詞である単語からストップワードを除いて、
# カウンターに記録する&コーパスに記録し、リターンで返す
#使用には、ストップワード関数を必要とする
def count_morpheme(word_class,mecab_file):

    copus=[]
    #語数カウンター
    word_counter = Counter()
    for line in make_lines(mecab_file):
        for morpheme in line:
            if morpheme['pos'] == word_class:
                if not len(morpheme['base'])==1:
                        word_counter.update([morpheme['base']])
                        #表層型をリスト型にする
                        copus.append(morpheme['base'])
                        copus.append(' ')
    copus=','.join(copus)
    return(copus,word_counter)


# あるエクセルの列に並んだ、単語を読みだす。
# そして順序指定型の辞書に保存する（語数と出現回数をセットで記録する）
def get_word(colum,path,sheet):
    #順序を保ったままの辞書型を設定する↓
    #(pythonの辞書型はデフォルトではソートされて挿入されない)
    word_dict=OrderedDict()
    #xlsxのパスを指定し、開く
    wb=openpyxl.load_workbook(path)
    #Sheetを指定し、読み込む
    sheet=wb.get_sheet_by_name(sheet)
    # print(sheet.columns)
    #指定した列を取得する、要素を使うためにはvalueが必要
    for i in range(1,200):
        word_dict[sheet.cell(row=i,column=colum).value]=0


    # for cell_obj in list(sheet.columns)[colum]:
    #     # word_dict.update({cell_obj.value:0})
    #     word_dict[cell_obj.value]=0
    return word_dict

    # あるエクセルの列に並んだ議員の名前を読み出し、
    # リスト型にして出力する
def get_name(colum,path,sheet):
    name_list=[]
    #xlsxのパスを指定し、開く
    wb=openpyxl.load_workbook(path)
    #Sheetを指定し、読み込む
    sheet=wb.get_sheet_by_name(sheet)

    #指定した列を取得する、要素を使うためにはvalueが必要
    for cell_obj in list(sheet.columns)[colum]:
        name_list.append(cell_obj.value)
    return name_list

    #もう使わないかな
def get_speech(colum,path,sheet):
    speech_list=[]
    #xlsxのパスを指定し、開く
    wb=openpyxl.load_workbook(path)
    #Sheetを指定し、読み込む
    sheet=wb.get_sheet_by_name(sheet)

    #指定した列を取得する、要素を使うためにはvalueが必要
    for cell_obj in list(sheet.columns)[colum]:
        mecab=MeCab.Tagger('-d /var/lib/mecab/dic/mecab-ipadic-neologd')
        # '-d /var/lib/mecab/dic/mecab-ipadic-neologd'
        cell_obj.value= mecab.parse(str(cell_obj.value))
        speech_list.append(cell_obj.value)
    return speech_list

#リスト型に収められた複数の文章を一つずつ読み出し、
# 特長語が入っているかカウントするかんすう
def count_people(speechList,char_num):
    for speech in speechList:
        for (key ,value) in char_num.items():
            # print(key)
            # print(char_num[key])
            # print(str(speech))
            if str(key) in str(speech):
                char_num[key]+=1

    print(char_num)










